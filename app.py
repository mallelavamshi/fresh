import streamlit as st
import os
import requests
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import tempfile
from datetime import datetime
import uuid
import gdown
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.utils import get_column_letter
import weasyprint
import base64

# Configure page
st.set_page_config(page_title="Image Analysis App", layout="wide")

def get_image_files(uploaded_files):
    """Process uploaded files"""
    valid_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')
    image_files = []
    
    for uploaded_file in uploaded_files:
        if uploaded_file.name.lower().endswith(valid_extensions):
            # Save uploaded file to temporary location
            with tempfile.NamedTemporaryFile(delete=False, suffix=uploaded_file.name) as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                image_files.append(tmp_file.name)
    
    return image_files

def process_google_drive_link(link):
    """Process Google Drive link and download images"""
    temp_dir = tempfile.mkdtemp()
    downloaded_files = []

    try:
        if 'drive.google.com' in link:
            if 'folder' in link:
                folder_id = link.split('/')[-1]
                if '?' in folder_id:
                    folder_id = folder_id.split('?')[0]
                gdown.download_folder(url=link, output=temp_dir, quiet=False)
                downloaded_files = [os.path.join(temp_dir, f) for f in os.listdir(temp_dir) 
                                 if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'))]
            else:
                file_id = link.split('/')[-2]
                output = os.path.join(temp_dir, "downloaded_image.jpg")
                gdown.download(f"https://drive.google.com/uc?id={file_id}", output, quiet=False)
                if os.path.exists(output):
                    downloaded_files = [output]
    except Exception as e:
        st.error(f"Error downloading from Google Drive: {str(e)}")
    
    return temp_dir, downloaded_files

def upload_file(image_path, api_key):
    """Upload file to get upload_file_id"""
    upload_url = 'https://api.dify.ai/v1/files/upload'
    headers = {
        'Authorization': f'Bearer {api_key}'
    }

    with open(image_path, 'rb') as image_file:
        files = {
            'file': (os.path.basename(image_path), image_file, 'image/jpeg')
        }
        response = requests.post(upload_url, headers=headers, files=files)
        response.raise_for_status()
        return response.json().get('id')

def process_image(image_path, api_key):
    """Process single image through Dify AI API"""
    try:
        file_id = upload_file(image_path, api_key)

        url = 'https://api.dify.ai/v1/chat-messages'
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }

        payload = {
            "inputs": {},
            "query": "Analyze this image and provide a detailed description and value assessment",
            "response_mode": "streaming",
            "conversation_id": "",
            "user": "abc-123",
            "files": [
                {
                    "type": "image",
                    "transfer_method": "local_file",
                    "upload_file_id": file_id
                }
            ]
        }

        response = requests.post(url, headers=headers, json=payload, stream=True)
        response.raise_for_status()
        
        full_response = ""
        for line in response.iter_lines():
            if line:
                try:
                    line_text = line.decode('utf-8')
                    if line_text.startswith('data: '):
                        line_text = line_text[6:]
                    
                    data = json.loads(line_text)
                    
                    if data.get('event') == 'agent_message' and 'answer' in data:
                        full_response += data['answer']
                        
                except (json.JSONDecodeError, Exception):
                    continue
        
        return full_response

    except requests.exceptions.RequestException as e:
        return f"Error processing image: {str(e)}"

def create_excel_with_images(results):
    """Create Excel file with images and analysis"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Results"

    # Set column headers
    ws['A1'] = "Image"
    ws['B1'] = "Image Name"
    ws['C1'] = "Analysis"

    # Style headers
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30

    row = 2
    for result in results:
        analysis_text = result['API_Response']
        ws.cell(row=row, column=2, value=result['Image_Name'])
        analysis_cell = ws.cell(row=row, column=3, value=analysis_text)
        
        analysis_cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Calculate row height
        approx_chars_per_line = 30
        num_lines = len(analysis_text) / approx_chars_per_line
        text_height = max(75, min(400, num_lines * 15))
        row_height = max(text_height, 200)
        ws.row_dimensions[row].height = row_height * 0.75

        try:
            img = Image.open(result['Image_Path'])
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            
            # Resize image
            max_width = 200
            aspect_ratio = img.width / img.height
            new_width = min(max_width, img.width)
            new_height = int(min(row_height, new_width / aspect_ratio))
            
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            
            xl_img = XLImage(img_byte_arr)
            ws.add_image(xl_img, f'A{row}')
            
        except Exception as e:
            ws.cell(row=row, column=1, value="Error loading image")
        
        row += 1

    # Save to bytes buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def main():
    st.title("Image Analysis App")
    
    # Sidebar for API key
    api_key = st.sidebar.text_input("Enter your API key:", type="password")
    
    # Main content
    st.write("Choose input method:")
    choice = st.radio("Select input method:", ["Local files", "Google Drive link"])
    
    results = []
    
    if choice == "Local files":
        uploaded_files = st.file_uploader("Upload images", accept_multiple_files=True)
        if uploaded_files:
            image_files = get_image_files(uploaded_files)
            
            if st.button("Process Images"):
                if not api_key:
                    st.error("Please enter an API key")
                    return
                    
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for i, image_path in enumerate(image_files):
                    status_text.text(f"Processing image {i+1}/{len(image_files)}")
                    response = process_image(image_path, api_key)
                    
                    results.append({
                        'Image_Name': os.path.basename(image_path),
                        'Image_Path': image_path,
                        'API_Response': response
                    })
                    
                    progress_bar.progress((i + 1) / len(image_files))
                
                if results:
                    excel_buffer = create_excel_with_images(results)
                    
                    # Create download buttons
                    st.download_button(
                        label="Download Excel",
                        data=excel_buffer,
                        file_name="analysis_results.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Display results
                    st.success("Processing complete!")
                    
                    # Clean up temporary files
                    for image_path in image_files:
                        try:
                            os.remove(image_path)
                        except:
                            pass
    
    else:
        drive_link = st.text_input("Enter Google Drive link:")
        if drive_link and st.button("Process Images"):
            if not api_key:
                st.error("Please enter an API key")
                return
                
            temp_dir, image_files = process_google_drive_link(drive_link)
            
            if image_files:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for i, image_path in enumerate(image_files):
                    status_text.text(f"Processing image {i+1}/{len(image_files)}")
                    response = process_image(image_path, api_key)
                    
                    results.append({
                        'Image_Name': os.path.basename(image_path),
                        'Image_Path': image_path,
                        'API_Response': response
                    })
                    
                    progress_bar.progress((i + 1) / len(image_files))
                
                if results:
                    excel_buffer = create_excel_with_images(results)
                    
                    # Create download buttons
                    st.download_button(
                        label="Download Excel",
                        data=excel_buffer,
                        file_name="analysis_results.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Display results
                    st.success("Processing complete!")
                    
                    # Clean up
                    import shutil
                    shutil.rmtree(temp_dir)
            else:
                st.error("No images found in the Google Drive link")

if __name__ == "__main__":
    main()